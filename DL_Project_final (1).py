import warnings
warnings.filterwarnings("ignore")
import gc
import os
import pickle
import traceback
import numpy as np
import torch
import torchvision
import torchvision.transforms as transforms
from torchvision import models
from torch.utils.data import DataLoader, Subset
from captum.attr import IntegratedGradients, LayerGradCam
import shap
import lime
import lime.lime_image
from skimage.segmentation import slic
from scipy import stats
from scipy.stats import f_oneway, spearmanr, pearsonr, mannwhitneyu
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy
import time
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from collections import defaultdict

print("=" * 60)
print("XAI Analysis Pipeline: VGG-16, ResNet-50 & ViT-B16")
print("=" * 60)

# ── Device ────────────────────────────────────────────────────────────────────
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Device: {DEVICE}")

# F12 — log GPU memory at startup so OOM risk is visible immediately
if torch.cuda.is_available():
    free, total = torch.cuda.mem_get_info()
    print(f"GPU memory: {free/1e9:.1f} GB free / {total/1e9:.1f} GB total")

# ── Imagenette class mapping ───────────────────────────────────────────────────
IMAGENETTE_CLASSES = {
    0: "tench",         1: "English springer", 2: "cassette player",
    3: "chain saw",     4: "church",           5: "French horn",
    6: "garbage truck", 7: "gas pump",         8: "golf ball",
    9: "parachute"
}
IMAGENET_INDICES = [0, 217, 482, 491, 497, 566, 569, 571, 574, 701]

# ── Constants ──────────────────────────────────────────────────────────────────
IMG_SIZE    = 224
NUM_SAMPLES = 750    # LIME / KernelSHAP perturbation samples
IG_N_STEPS  = 75     # Integrated Gradients steps
STABILITY_N = 10     # noisy samples for stability metric
NUM_IMAGES  = 20     # total images  (2 per class, 10 classes)
METHODS     = ["GradCAM", "IntGrad", "LIME", "KernelSHAP"]

# AOPC: 10 deletion steps 5%→50%
AOPC_FRACS  = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.45, 0.50]
SUFF_FRACS  = (0.1, 0.2, 0.3)

# F9 — rolling checkpoint interval (images)
CKPT_EVERY  = 5

# ── Output paths ──────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
HEATMAP_DIR = os.path.join(BASE_DIR, "xai_heatmaps")
CKPT_PATH   = os.path.join(BASE_DIR, "xai_checkpoint.pkl")
os.makedirs(HEATMAP_DIR, exist_ok=True)

# ── Transforms ────────────────────────────────────────────────────────────────
transform = transforms.Compose([
    transforms.Resize((IMG_SIZE, IMG_SIZE)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225])
])

# ── Dataset ───────────────────────────────────────────────────────────────────
print("\n[1/7] Loading Imagenette dataset …")
try:
    full_dataset = torchvision.datasets.Imagenette(
        root="/tmp/imagenette6", split="val",
        size="320px", transform=transform, download=True
    )
    per_class = defaultdict(list)
    if hasattr(full_dataset, "samples"):
        for idx, (_, lbl) in enumerate(full_dataset.samples):
            per_class[lbl].append(idx)
    elif hasattr(full_dataset, "targets"):
        for idx, lbl in enumerate(full_dataset.targets):
            per_class[lbl].append(idx)
    else:
        for idx in range(len(full_dataset)):
            per_class[full_dataset[idx][1]].append(idx)

    images_per_class = max(1, NUM_IMAGES // 10)
    selected = []
    for cls in range(10):
        selected.extend(per_class[cls][:images_per_class])
    selected = selected[:NUM_IMAGES]
    dataset  = Subset(full_dataset, selected)
    print(f"     Loaded {len(dataset)} images "
          f"({images_per_class}/class, {len(per_class)} classes found)")

except Exception as e:
    print(f"     Download failed ({e}); using synthetic data …")
    class SyntheticDataset(torch.utils.data.Dataset):
        def __len__(self): return NUM_IMAGES
        def __getitem__(self, i):
            torch.manual_seed(i)
            return torch.randn(3, IMG_SIZE, IMG_SIZE), i % 10
    dataset = SyntheticDataset()

loader = DataLoader(dataset, batch_size=1, shuffle=False)

# ── Models ────────────────────────────────────────────────────────────────────
print("\n[2/7] Loading VGG-16, ResNet-50, ViT-B16 …")
vgg16    = models.vgg16(weights=models.VGG16_Weights.IMAGENET1K_V1).to(DEVICE).eval()
resnet50 = models.resnet50(weights=models.ResNet50_Weights.IMAGENET1K_V1).to(DEVICE).eval()
vit_b16  = models.vit_b_16(weights=models.ViT_B_16_Weights.IMAGENET1K_V1).to(DEVICE).eval()
print("     Models loaded ✓")

# ── Helpers ───────────────────────────────────────────────────────────────────
mean_t = torch.tensor([0.485, 0.456, 0.406]).view(3, 1, 1)
std_t  = torch.tensor([0.229, 0.224, 0.225]).view(3, 1, 1)

def denorm(t):
    return (t.cpu() * std_t + mean_t).clamp(0, 1)

def top_imagenette_class(model, img_tensor):
    """Restrict 1000-class softmax to the 10 Imagenette classes."""
    with torch.no_grad():
        logits = model(img_tensor.to(DEVICE))
    probs_10       = torch.softmax(logits[0, IMAGENET_INDICES], dim=0)
    imagenette_idx = int(probs_10.argmax().item())
    conf           = float(probs_10[imagenette_idx].item())
    imagenet_class = IMAGENET_INDICES[imagenette_idx]
    return imagenette_idx, imagenet_class, conf

def normalize_attr(attr):
    a = attr - attr.min()
    if a.max() > 0:
        a = a / a.max()
    return a

def cuda_cleanup():
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.synchronize()
        torch.cuda.empty_cache()

def _save_checkpoint(results, sanity_results, stored_attrs, heatmap_files,
                     path=CKPT_PATH):
    with open(path, "wb") as f:
        pickle.dump({
            "results":        dict(results),
            "sanity_results": dict(sanity_results),
            "stored_attrs":   {k: v.detach().clone()
                               for k, v in stored_attrs.items()},
            "heatmap_files":  heatmap_files,
        }, f)
    print(f"  ✓ Checkpoint saved → {path} "
          f"({os.path.getsize(path)/1e6:.1f} MB)")

# ── Heatmap ───────────────────────────────────────────────────────────────────
def save_heatmap(img_tensor, attr_maps, img_idx, model_name, class_name):
    img_np          = denorm(img_tensor.squeeze()).numpy().transpose(1, 2, 0)
    methods_in_plot = list(attr_maps.keys())
    n_cols          = 1 + len(methods_in_plot)
    fig, axes       = plt.subplots(1, n_cols, figsize=(4 * n_cols, 4))
    fig.suptitle(f"Image {img_idx+1} | {model_name} | Class: {class_name}",
                 fontsize=12, fontweight="bold")
    axes[0].imshow(img_np); axes[0].set_title("Original"); axes[0].axis("off")
    cmaps = {"GradCAM": "jet", "IntGrad": "hot",
             "LIME": "RdBu", "KernelSHAP": "seismic"}
    for ax, mn in zip(axes[1:], methods_in_plot):
        heat = attr_maps[mn].squeeze().detach().numpy()
        ax.imshow(img_np, alpha=0.5)
        im = ax.imshow(heat, cmap=cmaps.get(mn, "jet"),
                       alpha=0.6, vmin=heat.min(), vmax=heat.max())
        ax.set_title(mn); ax.axis("off")
        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    plt.tight_layout()
    fname = os.path.join(HEATMAP_DIR,
        f"img{img_idx+1:02d}_{model_name.replace('-','').lower()}.png")
    plt.savefig(fname, dpi=100, bbox_inches="tight")
    plt.close(fig)
    return fname


# ══════════════════════════════════════════════════════════════════════════════
# XAI Methods
# ══════════════════════════════════════════════════════════════════════════════

# ── KernelSHAP ────────────────────────────────────────────────────────────────
def compute_kernelshap(model, img_tensor, target_class, n_samples=NUM_SAMPLES):
    img_np   = denorm(img_tensor.squeeze()).numpy().transpose(1, 2, 0)
    segments = slic(img_np, n_segments=50, compactness=10,
                    sigma=1, start_label=0)
    n_segs   = segments.max() + 1
    grey_np  = np.full_like(img_np, 0.5, dtype=np.float32)
    CHUNK    = 32   # lower to 16 if GPU < 6 GB still hits OOM

    def predict_fn(segment_masks):
        all_outs = []
        for start in range(0, len(segment_masks), CHUNK):
            chunk = segment_masks[start: start + CHUNK]
            batch = []
            for row in chunk:
                comp = img_np.copy()
                for seg in range(n_segs):
                    if row[seg] < 0.5:   # < 0.5 handles float 0.0 from SHAP
                        comp[segments == seg] = grey_np[segments == seg]
                batch.append(comp)
            bt = torch.tensor(np.array(batch, dtype=np.float32)) \
                      .permute(0, 3, 1, 2).to(DEVICE)
            mn = torch.tensor([0.485,0.456,0.406]).view(1,3,1,1).to(DEVICE)
            sd = torch.tensor([0.229,0.224,0.225]).view(1,3,1,1).to(DEVICE)
            bt = (bt - mn) / sd
            with torch.no_grad():
                out = torch.softmax(model(bt), 1).cpu().numpy()
            all_outs.append(out)
            del bt, mn, sd
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
        return np.concatenate(all_outs, axis=0)

    cuda_cleanup()   # F4: flush GPU before the 750-iteration SHAP loop
    background  = np.zeros((1, n_segs))
    image_input = np.ones((1, n_segs))
    explainer   = shap.KernelExplainer(predict_fn, background)
    shap_vals   = explainer.shap_values(image_input,
                                        nsamples=n_samples, silent=True)

    if isinstance(shap_vals, list):
        sv = np.array(shap_vals[target_class]).flatten()
    else:
        sv = np.array(shap_vals).flatten()

    if   len(sv) > n_segs: sv = sv[:n_segs]
    elif len(sv) < n_segs: sv = np.pad(sv, (0, n_segs - len(sv)))

    heat = np.zeros(segments.shape, dtype=np.float32)
    for seg in range(n_segs):
        heat[segments == seg] = sv[seg]
    return normalize_attr(torch.tensor(heat, dtype=torch.float32).unsqueeze(0))


# ── Grad-CAM ──────────────────────────────────────────────────────────────────
def get_gradcam_layer(model, model_name):
    if "vgg" in model_name.lower(): return model.features[28]
    return model.layer4   # ResNet-50

def _compute_gradcam_vit(model, img_tensor, target_class):
    model.eval()   # F11: ensure eval mode throughout
    activations, gradients = [None], [None]

    def fwd_hook(module, inp, out):
        activations[0] = out.detach().clone()   # (1, 197, 768)

    def bwd_hook(module, grad_in, grad_out):
        gradients[0] = grad_out[0].detach().clone()

    layer = model.encoder.layers[-1]
    fh = layer.register_forward_hook(fwd_hook)
    bh = layer.register_full_backward_hook(bwd_hook)

    img = img_tensor.to(DEVICE)
    try:
        with torch.enable_grad():
            model.zero_grad()
            model(img)[0, target_class].backward()
        if torch.cuda.is_available():
            torch.cuda.synchronize()   # F3: catch async CUDA errors here
    finally:
        fh.remove(); bh.remove()
        model.zero_grad()   # clear grad state before returning

    act = activations[0][:, 1:, :]   # strip CLS → (1, 196, 768)
    grd = gradients[0][:, 1:, :]
    weights = grd.mean(dim=-1, keepdim=True)
    cam     = torch.relu((weights * act).sum(dim=-1))  # (1, 196)

    n = cam.shape[-1]
    h = w = int(n ** 0.5)
    assert h * w == n, f"ViT patch count {n} not a perfect square."

    cam = cam.reshape(1, 1, h, w)
    cam = torch.nn.functional.interpolate(
        cam.float(), size=(IMG_SIZE, IMG_SIZE),
        mode="bilinear", align_corners=False)
    return normalize_attr(cam.squeeze(0).cpu())

def compute_gradcam(model, img_tensor, target_class, model_name):
    if "vit" in model_name.lower():
        return _compute_gradcam_vit(model, img_tensor, target_class)
    layer = get_gradcam_layer(model, model_name)
    gc_   = LayerGradCam(model, layer)
    attr  = gc_.attribute(img_tensor.to(DEVICE), target=target_class)
    attr  = torch.nn.functional.interpolate(
        attr, size=(IMG_SIZE, IMG_SIZE), mode="bilinear", align_corners=False)
    return normalize_attr(attr.squeeze(0).mean(0, keepdim=True).cpu())


# ── Integrated Gradients ──────────────────────────────────────────────────────
def compute_ig(model, img_tensor, target_class):
    ig   = IntegratedGradients(model)
    base = torch.zeros_like(img_tensor).to(DEVICE)
    attr = ig.attribute(img_tensor.to(DEVICE), base,
                        target=target_class, n_steps=IG_N_STEPS)
    return normalize_attr(attr.squeeze(0).abs().mean(0, keepdim=True).cpu())


# ── LIME ──────────────────────────────────────────────────────────────────────
def compute_lime(model, img_tensor, target_class, n_samples=NUM_SAMPLES):
    img_np   = denorm(img_tensor.squeeze()).numpy().transpose(1, 2, 0)
    LIME_CHUNK = 64

    def predict_fn(imgs):
        arr      = np.array(imgs, dtype=np.float32)
        all_outs = []
        mn = torch.tensor([0.485,0.456,0.406]).view(1,3,1,1).to(DEVICE)
        sd = torch.tensor([0.229,0.224,0.225]).view(1,3,1,1).to(DEVICE)
        for start in range(0, len(arr), LIME_CHUNK):
            t = torch.tensor(arr[start:start+LIME_CHUNK]) \
                     .permute(0, 3, 1, 2).to(DEVICE)
            with torch.no_grad():
                all_outs.append(
                    torch.softmax(model((t - mn) / sd), 1).cpu().numpy())
        return np.concatenate(all_outs, axis=0)

    expl    = lime.lime_image.LimeImageExplainer()
    exp     = expl.explain_instance(img_np, predict_fn,
                                    labels=[target_class],
                                    num_samples=n_samples, top_labels=1)
    lbl     = exp.top_labels[0]
    _, mask = exp.get_image_and_mask(lbl, positive_only=False,
                                     num_features=20, hide_rest=False)
    return normalize_attr(torch.tensor(mask, dtype=torch.float32).unsqueeze(0))


# ══════════════════════════════════════════════════════════════════════════════
# Metrics
# ══════════════════════════════════════════════════════════════════════════════

def faithfulness(model, img_tensor, attr_map, target_class,
                 k_fracs=(0.1, 0.2, 0.3)):
    img = img_tensor.to(DEVICE)
    with torch.no_grad():
        base = torch.softmax(model(img), 1)[0, target_class].item()
    flat = attr_map.flatten(); drops = []
    for frac in k_fracs:
        k    = max(1, int(frac * flat.numel()))
        mask = torch.ones(flat.numel()); mask[flat.topk(k).indices] = 0
        m3   = mask.reshape(attr_map.shape).expand(3,-1,-1).unsqueeze(0).to(DEVICE)
        with torch.no_grad():
            drops.append(base - torch.softmax(model(img * m3), 1)[0, target_class].item())
    return float(np.mean(drops))


def aopc(model, img_tensor, attr_map, target_class, k_fracs=AOPC_FRACS):
    img = img_tensor.to(DEVICE)
    with torch.no_grad():
        base = torch.softmax(model(img), 1)[0, target_class].item()
    flat = attr_map.flatten(); si = flat.argsort(descending=True); drops = []
    for frac in k_fracs:
        k    = max(1, int(frac * flat.numel()))
        mask = torch.ones(flat.numel()); mask[si[:k]] = 0
        m3   = mask.reshape(attr_map.shape).expand(3,-1,-1).unsqueeze(0).to(DEVICE)
        with torch.no_grad():
            drops.append(base - torch.softmax(model(img * m3), 1)[0, target_class].item())
    return float(np.mean(drops))


def sufficiency(model, img_tensor, attr_map, target_class,
                k_fracs=SUFF_FRACS):
    img = img_tensor.to(DEVICE); flat = attr_map.flatten(); scores = []
    for frac in k_fracs:
        k    = max(1, int(frac * flat.numel()))
        mask = torch.zeros(flat.numel()); mask[flat.topk(k).indices] = 1
        m3   = mask.reshape(attr_map.shape).expand(3,-1,-1).unsqueeze(0).to(DEVICE)
        with torch.no_grad():
            scores.append(torch.softmax(model(img * m3), 1)[0, target_class].item())
    return float(np.mean(scores))


def sparsity(attr_map, threshold=0.5):
    return float((attr_map > threshold).float().mean().item())


def stability(model, img_tensor, attr_fn, target_class,
              noise_std=0.05, n=STABILITY_N):
    base = attr_fn(model, img_tensor, target_class).flatten().detach().numpy()
    sims = []
    for _ in range(n):
        noisy = img_tensor + torch.randn_like(img_tensor) * noise_std
        na    = attr_fn(model, noisy, target_class).flatten().detach().numpy()
        sims.append(np.dot(base, na) /
                    (np.linalg.norm(base) * np.linalg.norm(na) + 1e-8))
    return float(np.mean(sims))


# ── Sanity check ──────────────────────────────────────────────────────────────
def cascading_randomization(model, img_tensor, target_class,
                             attr_fn, model_name):
    img_tensor = img_tensor.to(DEVICE)
    mc         = copy.deepcopy(model).to(DEVICE)
    orig       = attr_fn(model, img_tensor.cpu(),
                         target_class).flatten().detach().numpy()
    params  = [(n, p) for n, p in mc.named_parameters() if p.requires_grad]
    subset  = params[::max(1, len(params) // 4)]
    out     = []
    for lname, param in subset:
        with torch.no_grad():
            param.data = torch.randn_like(param.data)
        ra = attr_fn(mc, img_tensor.cpu(),
                     target_class).flatten().detach().numpy()
        r, _ = stats.pearsonr(orig, ra)
        out.append((lname, float(r)))
    return out


# ── Attribution agreement ─────────────────────────────────────────────────────
def compute_attribution_agreement(stored_attrs):
    pairs     = [("VGG-16","ResNet-50"),("VGG-16","ViT-B16"),
                 ("ResNet-50","ViT-B16")]
    agreement = defaultdict(list)
    for img_idx in range(NUM_IMAGES):
        for method in METHODS:
            for m1, m2 in pairs:
                k1, k2 = (img_idx,m1,method), (img_idx,m2,method)
                if k1 not in stored_attrs or k2 not in stored_attrs:
                    continue
                a = stored_attrs[k1].flatten().detach().numpy()  # F1
                b = stored_attrs[k2].flatten().detach().numpy()  # F1
                r, _ = spearmanr(a, b)
                agreement[(method, f"{m1} vs {m2}")].append(float(r))
    return agreement


# ── Confidence stratification ─────────────────────────────────────────────────
def confidence_stratified_analysis(results_dict, threshold=0.90):
    high_conf, low_conf = defaultdict(list), defaultdict(list)
    for (mn, mth), records in results_dict.items():
        for rec in records:
            (high_conf if rec["base_conf"] >= threshold
             else low_conf)[(mn, mth)].append(rec)
    metrics = ["faithfulness","aopc","sufficiency","sparsity","stability"]
    result  = {}
    for key in high_conf:
        row = {}
        for m in metrics:
            hv = [r[m] for r in high_conf[key] if not np.isnan(r.get(m,np.nan))]
            lv = [r[m] for r in low_conf.get(key,[])
                  if not np.isnan(r.get(m,np.nan))]
            hm = float(np.mean(hv)) if hv else np.nan
            lm = float(np.mean(lv)) if lv else np.nan
            d  = (hm - lm) if not (np.isnan(hm) or np.isnan(lm)) else np.nan
            p  = float(mannwhitneyu(hv, lv, alternative="two-sided")[1]) \
                 if len(hv) >= 3 and len(lv) >= 3 else np.nan
            row[m] = {"hi_mean":hm,"lo_mean":lm,"delta":d,"p_val":p,
                      "n_high":len(hv),"n_low":len(lv)}
        result[key] = row
    return result


# ── Metric correlations ───────────────────────────────────────────────────────
def compute_metric_correlations(results_dict):
    names  = ["faithfulness","aopc","sufficiency","sparsity","stability"]
    pairs  = [(m1,m2) for i,m1 in enumerate(names) for m2 in names[i+1:]]
    pooled = {m: [] for m in names}
    for records in results_dict.values():
        for rec in records:
            for m in names:
                v = rec.get(m, np.nan)
                if not np.isnan(v): pooled[m].append(v)
    global_corrs = {}
    for m1, m2 in pairs:
        v1, v2 = np.array(pooled[m1]), np.array(pooled[m2])
        ok = ~(np.isnan(v1)|np.isnan(v2))
        if ok.sum() >= 5:
            r, p = pearsonr(v1[ok], v2[ok])
            global_corrs[(m1,m2)] = (float(r), float(p))
        else:
            global_corrs[(m1,m2)] = (np.nan, np.nan)
    per_key = {}
    for (mn,mth), records in results_dict.items():
        kc = {}
        for m1, m2 in pairs:
            v1 = np.array([rec.get(m1,np.nan) for rec in records])
            v2 = np.array([rec.get(m2,np.nan) for rec in records])
            ok = ~(np.isnan(v1)|np.isnan(v2))
            if ok.sum() >= 5:
                r, p = pearsonr(v1[ok], v2[ok])
                kc[(m1,m2)] = (float(r), float(p))
            else:
                kc[(m1,m2)] = (np.nan, np.nan)
        per_key[(mn,mth)] = kc
    return global_corrs, per_key


# ══════════════════════════════════════════════════════════════════════════════
# Main Loop
# ══════════════════════════════════════════════════════════════════════════════
print("\n[3/7] Running explanations and computing metrics …")

results        = defaultdict(list)
sanity_results = defaultdict(list)
heatmap_files  = []
stored_attrs   = {}   # {(img_idx, model_name, method): detached attr}

models_dict = {"VGG-16": vgg16, "ResNet-50": resnet50, "ViT-B16": vit_b16}

for img_idx, (img_tensor, label) in enumerate(loader):
    label_int  = int(label.item())
    img_tensor = img_tensor.cpu()
    print(f"\n  Image {img_idx+1}/{len(dataset)}  "
          f"(label: {IMAGENETTE_CLASSES.get(label_int, label_int)})")

    for model_name, model in models_dict.items():
        print(f"    Model: {model_name}")

        # F8: initialize defaults so except block is safe even if
        #     top_imagenette_class fails (OOM, etc.)
        imagenette_idx, target_class, base_conf = 0, 0, 0.0
        try:
            imagenette_idx, target_class, base_conf = \
                top_imagenette_class(model, img_tensor)
        except Exception as e:
            print(f"      [SKIP] top_imagenette_class failed: {e}")
            cuda_cleanup()
            continue   # skip this model for this image

        predicted_name = IMAGENETTE_CLASSES[imagenette_idx]
        print(f"      → Predicted: {predicted_name} "
              f"(idx {target_class}, conf {base_conf:.3f})")

        attr_maps_for_plot = {}

        for method_name in METHODS:
            print(f"      Method: {method_name} …", end=" ", flush=True)
            t0 = time.time()
            try:
                # ── Compute attribution ───────────────────────────────────
                if method_name == "GradCAM":
                    attr = compute_gradcam(model, img_tensor,
                                           target_class, model_name)
                    def _attr_fn(m, img, tc, mn=model_name):
                        return compute_gradcam(m, img, tc, mn)
                elif method_name == "IntGrad":
                    attr     = compute_ig(model, img_tensor, target_class)
                    _attr_fn = compute_ig
                elif method_name == "LIME":
                    attr     = compute_lime(model, img_tensor, target_class)
                    _attr_fn = compute_lime
                else:   # KernelSHAP
                    attr     = compute_kernelshap(model, img_tensor, target_class)
                    _attr_fn = compute_kernelshap

                # F2: detach at storage — ViT backward leaves requires_grad=True
                stored_attrs[(img_idx, model_name, method_name)] = \
                    attr.detach().clone()

                attr_maps_for_plot[method_name] = attr

                faith = faithfulness(model, img_tensor, attr, target_class)
                ao    = aopc(model, img_tensor, attr, target_class)
                suff  = sufficiency(model, img_tensor, attr, target_class)
                spar  = sparsity(attr)
                stab  = stability(model, img_tensor, _attr_fn, target_class)

                print(f"faith={faith:.3f} aopc={ao:.3f} suff={suff:.3f} "
                      f"spar={spar:.3f} stab={stab:.3f} "
                      f"({time.time()-t0:.1f}s)")

                results[(model_name, method_name)].append({
                    "img": img_idx, "target_class": imagenette_idx,
                    "base_conf": base_conf,
                    "faithfulness": faith, "aopc": ao,
                    "sufficiency": suff, "sparsity": spar, "stability": stab,
                })

                if img_idx == 0:
                    print(f"        Sanity check …", end=" ", flush=True)
                    sanity_results[(model_name, method_name)] = \
                        cascading_randomization(model, img_tensor,
                                                target_class, _attr_fn,
                                                model_name)
                    print("done")

            except Exception as e:
                traceback.print_exc()
                print(f"ERROR: {e}")
                results[(model_name, method_name)].append({
                    "img": img_idx, "target_class": imagenette_idx,
                    "base_conf": base_conf,
                    "faithfulness": np.nan, "aopc": np.nan,
                    "sufficiency": np.nan, "sparsity": np.nan,
                    "stability": np.nan,
                })

            # F3: synchronize + empty_cache after every method
            cuda_cleanup()

        if attr_maps_for_plot:
            fpath = save_heatmap(img_tensor, attr_maps_for_plot,
                                 img_idx, model_name, predicted_name)
            heatmap_files.append(fpath)
            print(f"      ✓ Heatmap → {os.path.basename(fpath)}")

    # F9: rolling checkpoint every CKPT_EVERY images
    if (img_idx + 1) % CKPT_EVERY == 0 or (img_idx + 1) == NUM_IMAGES:
        print(f"  [Checkpoint] after image {img_idx+1}/{NUM_IMAGES} …",
              end=" ", flush=True)
        _save_checkpoint(results, sanity_results, stored_attrs, heatmap_files)

print(f"\n  ✓ {len(heatmap_files)} heatmaps saved → {HEATMAP_DIR}")

# F6: final checkpoint (ensures last image is always captured)
print("\n  Final checkpoint …", end=" ", flush=True)
_save_checkpoint(results, sanity_results, stored_attrs, heatmap_files)


# ══════════════════════════════════════════════════════════════════════════════
# Post-loop Analyses
# ══════════════════════════════════════════════════════════════════════════════
print("\n[4/7] Aggregating and running post-loop analyses …")

def agg(lst, key):
    vals = [d[key] for d in lst if not np.isnan(d.get(key, np.nan))]
    return (float(np.mean(vals)), float(np.std(vals))) if vals else (np.nan, np.nan)

summary = {
    key: {m: agg(recs, m) for m in
          ["faithfulness","aopc","sufficiency","sparsity","stability"]}
    for key, recs in results.items()
}

agreement_results           = compute_attribution_agreement(stored_attrs)
conf_strat                  = confidence_stratified_analysis(results)
global_corrs, per_key_corrs = compute_metric_correlations(results)
print(f"     Agreement: {len(agreement_results)} pairs ✓  |  "
      f"Stratification: {len(conf_strat)} keys ✓  |  "
      f"Correlations: {len(global_corrs)} metric pairs ✓")


# ── ANOVA ─────────────────────────────────────────────────────────────────────
print("[5/7] Running ANOVA …")

def anova_by_factor(metric, factor_key):
    groups = {}
    for (mn, mth), recs in results.items():
        vals = [d[metric] for d in recs if not np.isnan(d.get(metric,np.nan))]
        k    = mn if factor_key == 0 else mth
        if vals: groups.setdefault(k, []).extend(vals)
    if len(groups) < 2: return np.nan, np.nan
    f, p = f_oneway(*groups.values())
    return float(f), float(p)

ALL_M = ["faithfulness","aopc","sufficiency","sparsity","stability"]
anova_results = {
    "By Method": {m: anova_by_factor(m, 1) for m in ALL_M},
    "By Model":  {m: anova_by_factor(m, 0) for m in ALL_M},
}


# ══════════════════════════════════════════════════════════════════════════════
# Excel Report
# ══════════════════════════════════════════════════════════════════════════════
print("[6/7] Building Excel report …")

wb = openpyxl.Workbook()

H_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
S_FILL  = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
A_FILL  = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
G_FILL  = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
W_FILL  = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
B_FILL  = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
S_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
T_FONT  = Font(name="Arial", bold=True, color="1F4E79", size=14)
BD_FONT = Font(name="Arial", bold=True, size=10)
BF      = Font(name="Arial", size=10)
CTR     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin    = Side(style="thin",   color="AAAAAA")
thick   = Side(style="medium", color="1F4E79")
TB      = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
TK      = Border(left=thick, right=thick, top=thick, bottom=thick)

def Hdr(ws, r, c, v, w=None):
    x = ws.cell(r, c, v); x.fill=H_FILL; x.font=H_FONT
    x.alignment=CTR; x.border=TB
    if w: ws.column_dimensions[get_column_letter(c)].width = w

def Shdr(ws, r, c, v):
    x = ws.cell(r, c, v); x.fill=S_FILL; x.font=S_FONT
    x.alignment=CTR; x.border=TB

def Cell(ws, r, c, v, bold=False, align=CTR, fill=None, fmt=None):
    x = ws.cell(r, c, v); x.font = BD_FONT if bold else BF
    x.alignment=align; x.border=TB
    if fill: x.fill=fill
    if fmt:  x.number_format=fmt
    return x

def TRow(ws, r, text, nc):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    x = ws.cell(r, 1, text); x.font=T_FONT; x.alignment=CTR; x.border=TK

def sfill(p):
    if np.isnan(p): return None
    return G_FILL if p<0.01 else (W_FILL if p<0.05 else B_FILL)

def fmtm(m):
    return "AOPC" if m=="aopc" else m.capitalize()


# ── Sheet 1: Overview ─────────────────────────────────────────────────────────
ws1 = wb.active; ws1.title="Overview"
ws1.sheet_view.showGridLines=False; ws1.row_dimensions[1].height=30
TRow(ws1,1,"XAI Analysis: VGG-16, ResNet-50 & ViT-B16 on Imagenette",6)
info=[
    ("Dataset",      "Imagenette (10-class ImageNet subset, stratified val split)"),
    ("Models",       "VGG-16, ResNet-50, ViT-B16 (ImageNet pretrained)"),
    ("XAI Methods",  "Grad-CAM, Integrated Gradients, LIME, KernelSHAP (real)"),
    ("Core Metrics", "Faithfulness, AOPC, Sufficiency, Sparsity, Stability"),
    ("New Analyses", "Attribution Agreement, Confidence Stratification, Metric Correlations"),
    ("Sanity Check", "Cascading Parameter Randomization"),
    ("Statistics",   "One-way ANOVA (by model and by method)"),
    ("Images",       f"{NUM_IMAGES} ({NUM_IMAGES//10}/class, 10 classes)"),
    ("IG n_steps",   str(IG_N_STEPS)),
    ("LIME/SHAP n",  str(NUM_SAMPLES)),
    ("Stability n",  str(STABILITY_N)),
    ("AOPC steps",   f"{len(AOPC_FRACS)} ({int(AOPC_FRACS[0]*100)}%–{int(AOPC_FRACS[-1]*100)}%)"),
    ("Checkpoint",   CKPT_PATH),
    ("Heatmaps",     HEATMAP_DIR),
    ("Date",         time.strftime("%Y-%m-%d %H:%M")),
]
for i,(k,v) in enumerate(info,4):
    Cell(ws1,i,1,k,bold=True,align=LFT)
    ws1.merge_cells(start_row=i,start_column=2,end_row=i,end_column=6)
    Cell(ws1,i,2,v,align=LFT)
ws1.column_dimensions["A"].width=22; ws1.column_dimensions["B"].width=70


# ── Sheet 2: Detailed Results ─────────────────────────────────────────────────
ws2=wb.create_sheet("Detailed Results"); ws2.sheet_view.showGridLines=False
TRow(ws2,1,"Per-Image Metric Results",10)
for c,(h,w) in enumerate(zip(
    ["Model","Method","Image #","Target Class","Base Conf",
     "Faithfulness","AOPC","Sufficiency","Sparsity","Stability"],
    [14,14,10,16,12,14,12,14,12,12]),1): Hdr(ws2,2,c,h,w)
row=3
for (mn,mth),recs in sorted(results.items()):
    for rec in recs:
        fl=A_FILL if row%2==0 else None
        cn=IMAGENETTE_CLASSES.get(rec["target_class"],rec["target_class"])
        for c,v in enumerate([mn,mth,rec["img"]+1,cn,rec["base_conf"],
            rec["faithfulness"],rec["aopc"],rec["sufficiency"],
            rec["sparsity"],rec["stability"]],1):
            Cell(ws2,row,c,round(v,4) if isinstance(v,float) else v,
                 align=CTR,fill=fl,fmt="0.000" if c>=5 else None)
        row+=1
ws2.freeze_panes="A3"


# ── Sheet 3: Summary Statistics ───────────────────────────────────────────────
ws3=wb.create_sheet("Summary Statistics"); ws3.sheet_view.showGridLines=False
TRow(ws3,1,"Mean ± Std per (Model, Method)",12)
for c,(h,w) in enumerate(zip(
    ["Model","Method","Faith Mean","Faith Std","AOPC Mean","AOPC Std",
     "Suff Mean","Suff Std","Spar Mean","Spar Std","Stab Mean","Stab Std"],
    [14,14,13,12,12,12,12,12,12,12,12,12]),1): Hdr(ws3,2,c,h,w)
row=3
for (mn,mth),met in sorted(summary.items()):
    fl=A_FILL if row%2==0 else None
    vals=[mn,mth]
    for m in ["faithfulness","aopc","sufficiency","sparsity","stability"]:
        a,s=met[m]; vals+=[a,s]
    for c,v in enumerate(vals,1):
        Cell(ws3,row,c,round(v,4) if isinstance(v,float) else v,
             align=CTR,fill=fl,fmt="0.0000" if c>=3 else None)
    row+=1
ws3.freeze_panes="A3"


# ── Sheet 4: Comparison Matrix ────────────────────────────────────────────────
ws4=wb.create_sheet("Comparison Matrix"); ws4.sheet_view.showGridLines=False
ml   = ["GradCAM","IntGrad","LIME","KernelSHAP"]
mdl  = sorted({mn for (mn,_) in summary})
sr=1
for metric in ["faithfulness","aopc","sufficiency","sparsity","stability"]:
    TRow(ws4,sr,f"Metric: {fmtm(metric)} (mean values)",len(ml)+2)
    Shdr(ws4,sr+1,1,"Model \\ Method")
    for j,m in enumerate(ml,2): Shdr(ws4,sr+1,j,m)
    ws4.column_dimensions["A"].width=14
    for j in range(2,len(ml)+2):
        ws4.column_dimensions[get_column_letter(j)].width=14
    for i,mn in enumerate(mdl):
        rr=sr+2+i; Cell(ws4,rr,1,mn,bold=True,align=CTR)
        rvs=[summary.get((mn,m),{}).get(metric,(np.nan,np.nan))[0] for m in ml]
        vld=[v for v in rvs if not np.isnan(v)]
        for j,(m,mv) in enumerate(zip(ml,rvs),2):
            cx=Cell(ws4,rr,j,round(mv,4) if not np.isnan(mv) else "N/A",
                    align=CTR,fmt="0.0000")
            if vld and not np.isnan(mv):
                if mv==max(vld): cx.fill=G_FILL
                elif mv==min(vld): cx.fill=B_FILL
    sr+=len(mdl)+4


# ── Sheet 5: ANOVA ────────────────────────────────────────────────────────────
ws5=wb.create_sheet("ANOVA Results"); ws5.sheet_view.showGridLines=False
TRow(ws5,1,"One-Way ANOVA: Model and Method Comparisons",5)
for c,(h,w) in enumerate(zip(
    ["Factor","Metric","F-statistic","p-value","Significant (p<0.05)"],
    [16,16,14,14,22]),1): Hdr(ws5,2,c,h,w)
row=3
for fac,fd in anova_results.items():
    for met,(fv,pv) in fd.items():
        sig="Yes ✓" if (not np.isnan(pv) and pv<0.05) else "No"
        pf=sfill(pv) if not np.isnan(pv) else None
        for c,v in enumerate([fac,fmtm(met),
            round(fv,4) if not np.isnan(fv) else "N/A",
            round(pv,4) if not np.isnan(pv) else "N/A",sig],1):
            Cell(ws5,row,c,v,align=CTR,
                 fill=pf if c==5 else (A_FILL if row%2==0 else None))
        row+=1
row+=1; Cell(ws5,row,1,"Legend:",bold=True,align=LFT); row+=1
for fl,lb in [(G_FILL,"p < 0.01 — Highly significant"),
              (W_FILL,"p < 0.05 — Significant"),
              (B_FILL,"p ≥ 0.05 — Not significant")]:
    c=ws5.cell(row,1,lb); c.fill=fl; c.font=BF; c.alignment=LFT
    ws5.merge_cells(start_row=row,start_column=1,end_row=row,end_column=3)
    row+=1


# ── Sheet 6: Sanity Checks ────────────────────────────────────────────────────
ws6=wb.create_sheet("Sanity Checks"); ws6.sheet_view.showGridLines=False
TRow(ws6,1,
     "Cascading Parameter Randomization — Pearson r with Original Attribution",5)
for c,(h,w) in enumerate(zip(
    ["Model","Method","Layer Randomized","Pearson r","Interpretation"],
    [14,14,28,12,30]),1): Hdr(ws6,2,c,h,w)
row=3
for (mn,mth),sc in sorted(sanity_results.items()):
    for lname,rv in sc:
        fl=A_FILL if row%2==0 else None
        if not np.isnan(rv):
            interp=("High sensitivity (good)" if rv<0.3
                    else "Moderate sensitivity" if rv<0.7
                    else "Low sensitivity (poor)")
            cf=G_FILL if rv<0.3 else (W_FILL if rv<0.7 else B_FILL)
        else: interp,cf="N/A",None
        for c,v in enumerate([mn,mth,lname,
            round(rv,4) if not np.isnan(rv) else "N/A",interp],1):
            Cell(ws6,row,c,v,align=LFT if c==3 else CTR,
                 fill=cf if c==5 else fl)
        row+=1
ws6.column_dimensions["C"].width=30; ws6.column_dimensions["E"].width=30
ws6.freeze_panes="A3"


# ── Sheet 7: Sanity + Metrics ─────────────────────────────────────────────────
ws7=wb.create_sheet("Sanity + Metrics"); ws7.sheet_view.showGridLines=False
TRow(ws7,1,"Method Comparison Augmented with Sanity Check Scores",9)
for c,(h,w) in enumerate(zip(
    ["Model","Method","Avg Faith","Avg AOPC","Avg Suff",
     "Avg Spar","Avg Stab","Sanity Score (avg Pearson r)","Overall Score"],
    [14,14,14,12,12,12,12,26,14]),1): Hdr(ws7,2,c,h,w)
row=3
for (mn,mth) in sorted(summary):
    s=summary[(mn,mth)]
    fm,am,sfm,spm,stm=(s[x][0] for x in
        ["faithfulness","aopc","sufficiency","sparsity","stability"])
    sc=[r for _,r in sanity_results.get((mn,mth),[])]
    ss=float(np.mean(sc)) if sc else np.nan
    iv=(1-ss) if not np.isnan(ss) else np.nan
    ov_vals=[v for v in [fm,am,sfm,stm,iv] if not np.isnan(v)]
    ov=float(np.mean(ov_vals)) if ov_vals else np.nan
    fl=A_FILL if row%2==0 else None
    for c,v in enumerate([mn,mth,fm,am,sfm,spm,stm,ss,ov],1):
        Cell(ws7,row,c,
             round(v,4) if isinstance(v,float) and not np.isnan(v)
             else ("N/A" if isinstance(v,float) else v),
             align=CTR,fill=fl,fmt="0.0000" if c>=3 else None)
    row+=1
ws7.freeze_panes="A3"


# ── Sheet 8: Attribution Agreement ───────────────────────────────────────────
ws8=wb.create_sheet("Attribution Agreement"); ws8.sheet_view.showGridLines=False
TRow(ws8,1,
     "Inter-Architecture Attribution Agreement — Spearman r "
     "(same method, different models)",6)
for c,(h,w) in enumerate(zip(
    ["Method","Model Pair","Mean Spearman r","Std Dev","N Images","Interpretation"],
    [14,22,16,12,10,32]),1): Hdr(ws8,2,c,h,w)
row=3
for (mth,pair),rvs in sorted(agreement_results.items()):
    fl=A_FILL if row%2==0 else None
    vld=[v for v in rvs if not np.isnan(v)]
    if vld:
        mr=float(np.mean(vld)); sr2=float(np.std(vld))
        interp=("High agreement — input-driven" if mr>=0.7
                else "Moderate agreement" if mr>=0.4
                else "Low agreement — model-driven")
        cf=G_FILL if mr>=0.7 else (W_FILL if mr>=0.4 else B_FILL)
    else: mr,sr2,interp,cf=np.nan,np.nan,"N/A",None
    for c,v in enumerate([mth,pair,
        round(mr,4) if not np.isnan(mr) else "N/A",
        round(sr2,4) if not np.isnan(sr2) else "N/A",
        len(vld),interp],1):
        Cell(ws8,row,c,v,align=LFT if c==6 else CTR,
             fill=cf if c==6 else fl,
             fmt="0.0000" if c in(3,4) else None)
    row+=1
ws8.freeze_panes="A3"


# ── Sheet 9: Confidence Stratification ───────────────────────────────────────
ws9=wb.create_sheet("Confidence Stratification"); ws9.sheet_view.showGridLines=False
TRow(ws9,1,
     "High-Confidence vs Low-Confidence Predictions — "
     "Metric Comparison (threshold = 0.90)",9)
for c,(h,w) in enumerate(zip(
    ["Model","Method","Metric","High Conf Mean","Low Conf Mean",
     "Delta (Hi−Lo)","N High","N Low","p-value (MWU)"],
    [14,14,16,16,16,14,8,8,16]),1): Hdr(ws9,2,c,h,w)
row=3
for (mn,mth),mrows in sorted(conf_strat.items()):
    for met,vals in mrows.items():
        fl=A_FILL if row%2==0 else None
        pv=vals["p_val"]; pf=sfill(pv) if not np.isnan(pv) else None
        for c,v in enumerate([mn,mth,fmtm(met),
            round(vals["hi_mean"],4) if not np.isnan(vals["hi_mean"]) else "N/A",
            round(vals["lo_mean"],4) if not np.isnan(vals["lo_mean"]) else "N/A",
            round(vals["delta"],4)   if not np.isnan(vals["delta"])   else "N/A",
            vals["n_high"],vals["n_low"],
            round(pv,4) if not np.isnan(pv) else "N/A"],1):
            Cell(ws9,row,c,v,align=CTR,
                 fill=pf if c==9 else fl,
                 fmt="0.0000" if c in(4,5,6,9) else None)
        row+=1
ws9.freeze_panes="A3"


# ── Sheet 10: Metric Correlations ─────────────────────────────────────────────
ws10=wb.create_sheet("Metric Correlations"); ws10.sheet_view.showGridLines=False
TRow(ws10,1,
     "Pairwise Pearson r — Tests Whether Metrics Are Independent Dimensions",5)

ws10.merge_cells(start_row=2,start_column=1,end_row=2,end_column=5)
cx=ws10.cell(2,1,"Global Correlations (pooled across all images, models, methods)")
cx.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
cx.fill=S_FILL; cx.alignment=CTR
for c,(h,w) in enumerate(zip(
    ["Metric 1","Metric 2","Pearson r","p-value","Interpretation"],
    [16,16,12,12,32]),1): Hdr(ws10,3,c,h,w)
row=4
for (m1,m2),(rv,pv) in sorted(global_corrs.items()):
    fl=A_FILL if row%2==0 else None
    if not np.isnan(rv):
        ar=abs(rv)
        interp=("Strong — possibly redundant" if ar>=0.7
                else "Moderate" if ar>=0.4 else "Weak — independent")
        cf=B_FILL if ar>=0.7 else (W_FILL if ar>=0.4 else G_FILL)
    else: interp,cf="N/A",None
    for c,v in enumerate([fmtm(m1),fmtm(m2),
        round(rv,4) if not np.isnan(rv) else "N/A",
        round(pv,4) if not np.isnan(pv) else "N/A",interp],1):
        Cell(ws10,row,c,v,align=LFT if c==5 else CTR,
             fill=cf if c==5 else fl,fmt="0.0000" if c in(3,4) else None)
    row+=1

row+=2
ws10.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
cx=ws10.cell(row,1,"Per-(Model, Method) Correlations")
cx.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
cx.fill=S_FILL; cx.alignment=CTR; row+=1
for c,(h,w) in enumerate(zip(
    ["Model","Method","Metric Pair","Pearson r","p-value"],
    [14,14,26,12,12]),1): Hdr(ws10,row,c,h,w)
row+=1
for (mn,mth),kc in sorted(per_key_corrs.items()):
    for (m1,m2),(rv,pv) in sorted(kc.items()):
        fl=A_FILL if row%2==0 else None
        for c,v in enumerate([mn,mth,f"{fmtm(m1)} vs {fmtm(m2)}",
            round(rv,4) if not np.isnan(rv) else "N/A",
            round(pv,4) if not np.isnan(pv) else "N/A"],1):
            Cell(ws10,row,c,v,align=CTR,fill=fl,
                 fmt="0.0000" if c in(4,5) else None)
        row+=1
ws10.column_dimensions["C"].width=28
ws10.freeze_panes="A4"


# ── Save Excel ─────────────────────────────────────────────────────────────────
OUTPUT = os.path.join(BASE_DIR, "xai_analysis_report.xlsx")
wb.save(OUTPUT)
print(f"\n[7/7] Excel  → {OUTPUT}")
print(f"      Heatmaps  → {HEATMAP_DIR}/ ({len(heatmap_files)} files)")
print(f"      Checkpoint → {CKPT_PATH}")
print("\n      Sheets:")
for i,n in enumerate(["Overview","Detailed Results","Summary Statistics",
    "Comparison Matrix","ANOVA Results","Sanity Checks","Sanity + Metrics",
    "Attribution Agreement","Confidence Stratification","Metric Correlations"],1):
    print(f"        {i:2d}. {n}")
print("\n✅ Pipeline complete!")
